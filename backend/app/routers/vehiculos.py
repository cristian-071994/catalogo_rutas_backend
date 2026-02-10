from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from pydantic import BaseModel
from io import BytesIO
from openpyxl import load_workbook
from openpyxl import Workbook

from app.database.session import get_db
from app.models.vehiculo import Vehiculo
from app.models.configuracion_vehiculo import ConfiguracionVehiculo
from app.models.marca_vehiculo import MarcaVehiculo
from app.models.usuario import Usuario
from app.models.enums import EstadoGeneral
from app.schemas.vehiculo import VehiculoResponse
from app.auth import get_current_user, require_permission

router = APIRouter(
    prefix="/vehiculos",
    tags=["Vehículos"]
)


def _leer_excel_vehiculos(archivo: UploadFile) -> tuple[list[tuple], int, int, int]:
    if not archivo.filename or not archivo.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Archivo inválido. Debe ser .xlsx"
        )

    contenido = archivo.file.read()
    workbook = load_workbook(BytesIO(contenido))
    worksheet = workbook.active

    filas = list(worksheet.iter_rows(values_only=True))
    if not filas:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El archivo está vacío"
        )

    encabezados = [str(celda).strip().upper() if celda is not None else "" for celda in filas[0]]
    requeridos = ["PLACA", "MARCA", "MODELO"]
    faltantes = [col for col in requeridos if col not in encabezados]
    if faltantes:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Faltan columnas requeridas: {', '.join(faltantes)}"
        )

    idx_placa = encabezados.index("PLACA")
    idx_marca = encabezados.index("MARCA")
    idx_modelo = encabezados.index("MODELO")

    return filas, idx_placa, idx_marca, idx_modelo


class VehiculoCreate(BaseModel):
    """Para crear un vehículo"""
    placa: str
    configuracion_id: int


class VehiculoUpdate(BaseModel):
    """Para actualizar un vehículo"""
    placa: str = None
    configuracion_id: int = None
    estado: str = None


# ============================================
# OBTENER VEHÍCULOS
# ============================================

@router.get("/", response_model=list[VehiculoResponse], summary="Listar Vehículos")
def listar_vehiculos(
    incluir_inactivos: bool = False,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Lista todos los vehículos activos del sistema.
    
    GET /vehiculos/
    GET /vehiculos/?incluir_inactivos=true  <- Para incluir inactivos (soporte)
    
    Por defecto solo devuelve estado="activo".
    Agregue ?incluir_inactivos=true solo si necesita ver registros eliminados.
    """
    query = db.query(Vehiculo)
    
    # Multi-tenancy: filtrar por empresa excepto super_admin
    if current_user.rol and current_user.rol.nombre != "super_admin":
        query = query.filter(Vehiculo.empresa_id == current_user.empresa_id)
    
    if not incluir_inactivos:
        query = query.filter(Vehiculo.estado == EstadoGeneral.activo)
    
    return query.all()


@router.get("/{vehiculo_id}", response_model=VehiculoResponse, summary="Obtener Vehículo")
def obtener_vehiculo(
    vehiculo_id: int,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Obtiene un vehículo específico por su ID.
    
    GET /vehiculos/1
    """
    vehiculo = db.query(Vehiculo).filter(Vehiculo.id == vehiculo_id).first()

    if not vehiculo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Vehículo no encontrado"
        )

    # Multi-tenancy: verificar pertenencia a empresa excepto super_admin
    if current_user.rol and current_user.rol.nombre != "super_admin":
        if vehiculo.empresa_id != current_user.empresa_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="No tienes permiso para acceder a este recurso"
            )

    return vehiculo


# ============================================
# CREAR VEHÍCULO
# ============================================

@router.post(
    "/",
    response_model=VehiculoResponse,
    status_code=status.HTTP_201_CREATED
)
def crear_vehiculo(
    vehiculo: VehiculoCreate,
    current_user: Usuario = Depends(require_permission("crear_vehiculo")),
    db: Session = Depends(get_db)
):
    """
    Crea un nuevo vehículo.
    
    POST /vehiculos/
    Body:
    {
        "placa": "ABC123",
        "configuracion_id": 1
    }
    
    Esto significa:
    - Vehículo con placa ABC123
    - Es un Chevrolet 2020 (configuracion_id 1)
    
    ⚠️ La placa debe ser ÚNICA
    """
    
    # Validar que la configuración exista
    config = db.query(ConfiguracionVehiculo).filter(
        ConfiguracionVehiculo.id == vehiculo.configuracion_id
    ).first()

    if not config:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Configuración no encontrada"
        )

    # Validar que no exista con la misma placa (case-insensitive)
    existente = db.query(Vehiculo).filter(
        func.lower(Vehiculo.placa) == func.lower(vehiculo.placa)
    ).first()

    if existente:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Ya existe vehículo con placa '{vehiculo.placa}'"
        )

    # Crear
    nuevo_vehiculo = Vehiculo(
        **vehiculo.model_dump(),
        empresa_id=current_user.empresa_id  # Multi-tenancy: asignar empresa del usuario
    )
    db.add(nuevo_vehiculo)
    db.commit()
    db.refresh(nuevo_vehiculo)

    return nuevo_vehiculo


# ============================================
# CARGA MASIVA DE VEHICULOS (EXCEL)
# ============================================

@router.post("/carga-masiva", summary="Carga Masiva de Vehículos")
def carga_masiva_vehiculos(
    archivo: UploadFile = File(...),
    current_user: Usuario = Depends(require_permission("crear_vehiculo")),
    db: Session = Depends(get_db)
):
    """
    Carga masiva de vehículos desde archivo Excel (.xlsx).

    Columnas requeridas (en MAYUSCULA): PLACA, MARCA, MODELO
    """
    filas, idx_placa, idx_marca, idx_modelo = _leer_excel_vehiculos(archivo)

    marcas_cache: dict[str, MarcaVehiculo] = {}
    configs_cache: dict[tuple[int, int], ConfiguracionVehiculo] = {}
    placas_vistas: set[str] = set()

    creados = 0
    omitidos = []
    errores = []

    for row_idx, fila in enumerate(filas[1:], start=2):
        if not fila or all(celda is None for celda in fila):
            continue

        raw_placa = fila[idx_placa] if idx_placa < len(fila) else None
        raw_marca = fila[idx_marca] if idx_marca < len(fila) else None
        raw_modelo = fila[idx_modelo] if idx_modelo < len(fila) else None

        placa = str(raw_placa).strip().upper() if raw_placa else ""
        marca_nombre = str(raw_marca).strip() if raw_marca else ""

        try:
            modelo = int(raw_modelo) if raw_modelo is not None else None
        except (TypeError, ValueError):
            modelo = None

        if not placa or not marca_nombre or not modelo:
            errores.append({
                "fila": row_idx,
                "mensaje": "PLACA, MARCA y MODELO son obligatorios"
            })
            continue

        if placa in placas_vistas:
            omitidos.append({
                "fila": row_idx,
                "placa": placa,
                "mensaje": "Placa duplicada en el archivo"
            })
            continue
        placas_vistas.add(placa)

        existente = db.query(Vehiculo).filter(
            func.lower(Vehiculo.placa) == placa.lower()
        ).first()
        if existente:
            omitidos.append({
                "fila": row_idx,
                "placa": placa,
                "mensaje": "La placa ya existe"
            })
            continue

        marca_key = marca_nombre.strip().lower()
        marca = marcas_cache.get(marca_key)
        if not marca:
            marca = db.query(MarcaVehiculo).filter(
                func.lower(MarcaVehiculo.nombre) == marca_key
            ).first()
            if not marca:
                marca = MarcaVehiculo(nombre=marca_nombre.title())
                db.add(marca)
                db.flush()
            marcas_cache[marca_key] = marca

        config_key = (marca.id, modelo)
        config = configs_cache.get(config_key)
        if not config:
            config = db.query(ConfiguracionVehiculo).filter(
                ConfiguracionVehiculo.marca_id == marca.id,
                ConfiguracionVehiculo.modelo == modelo
            ).first()
            if not config:
                config = ConfiguracionVehiculo(marca_id=marca.id, modelo=modelo)
                db.add(config)
                db.flush()
            configs_cache[config_key] = config

        nuevo_vehiculo = Vehiculo(
            placa=placa,
            configuracion_id=config.id,
            empresa_id=current_user.empresa_id
        )
        db.add(nuevo_vehiculo)
        creados += 1

    db.commit()

    return {
        "creados": creados,
        "omitidos": omitidos,
        "errores": errores,
        "total": len(filas) - 1
    }


@router.post("/carga-masiva/preview", summary="Vista Previa Carga Masiva")
def vista_previa_carga_masiva(
    archivo: UploadFile = File(...),
    current_user: Usuario = Depends(require_permission("crear_vehiculo")),
    db: Session = Depends(get_db)
):
    """
    Vista previa de carga masiva. No crea datos, solo informa.
    """
    filas, idx_placa, idx_marca, idx_modelo = _leer_excel_vehiculos(archivo)

    marcas_cache: dict[str, MarcaVehiculo | None] = {}
    configs_cache: dict[tuple[int, int], ConfiguracionVehiculo | None] = {}
    placas_vistas: set[str] = set()
    nuevas_marcas: set[str] = set()
    nuevas_configs: set[tuple[int, int]] = set()

    items = []
    crear = 0
    omitidos = 0
    errores = 0

    for row_idx, fila in enumerate(filas[1:], start=2):
        if not fila or all(celda is None for celda in fila):
            continue

        raw_placa = fila[idx_placa] if idx_placa < len(fila) else None
        raw_marca = fila[idx_marca] if idx_marca < len(fila) else None
        raw_modelo = fila[idx_modelo] if idx_modelo < len(fila) else None

        placa = str(raw_placa).strip().upper() if raw_placa else ""
        marca_nombre = str(raw_marca).strip() if raw_marca else ""

        try:
            modelo = int(raw_modelo) if raw_modelo is not None else None
        except (TypeError, ValueError):
            modelo = None

        if not placa or not marca_nombre or not modelo:
            errores += 1
            items.append({
                "fila": row_idx,
                "placa": placa,
                "marca": marca_nombre,
                "modelo": raw_modelo,
                "accion": "error",
                "mensaje": "PLACA, MARCA y MODELO son obligatorios"
            })
            continue

        if placa in placas_vistas:
            omitidos += 1
            items.append({
                "fila": row_idx,
                "placa": placa,
                "marca": marca_nombre,
                "modelo": modelo,
                "accion": "omitir",
                "mensaje": "Placa duplicada en el archivo"
            })
            continue
        placas_vistas.add(placa)

        existente = db.query(Vehiculo).filter(
            func.lower(Vehiculo.placa) == placa.lower()
        ).first()
        if existente:
            omitidos += 1
            items.append({
                "fila": row_idx,
                "placa": placa,
                "marca": marca_nombre,
                "modelo": modelo,
                "accion": "omitir",
                "mensaje": "La placa ya existe"
            })
            continue

        marca_key = marca_nombre.strip().lower()
        marca = marcas_cache.get(marca_key)
        if marca is None:
            marca = db.query(MarcaVehiculo).filter(
                func.lower(MarcaVehiculo.nombre) == marca_key
            ).first()
            marcas_cache[marca_key] = marca

        nueva_marca = False
        if not marca:
            nueva_marca = True
            nuevas_marcas.add(marca_key)

        marca_id = marca.id if marca else -1
        config_key = (marca_id, modelo)
        config = configs_cache.get(config_key)
        if config is None and marca:
            config = db.query(ConfiguracionVehiculo).filter(
                ConfiguracionVehiculo.marca_id == marca.id,
                ConfiguracionVehiculo.modelo == modelo
            ).first()
            configs_cache[config_key] = config

        nueva_config = False
        if marca and not config:
            nueva_config = True
            nuevas_configs.add((marca.id, modelo))

        crear += 1
        items.append({
            "fila": row_idx,
            "placa": placa,
            "marca": marca_nombre,
            "modelo": modelo,
            "accion": "crear",
            "mensaje": "Se creara el vehiculo",
            "nueva_marca": nueva_marca,
            "nueva_config": nueva_config
        })

    return {
        "total": len(filas) - 1,
        "crear": crear,
        "omitidos": omitidos,
        "errores": errores,
        "nuevas_marcas": len(nuevas_marcas),
        "nuevas_configuraciones": len(nuevas_configs),
        "items": items
    }


@router.get("/carga-masiva/template", summary="Descargar Plantilla de Carga Masiva")
def descargar_plantilla_carga_masiva(
    current_user: Usuario = Depends(get_current_user)
):
    """
    Descarga una plantilla Excel con columnas PLACA, MARCA, MODELO.
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "vehiculos"
    worksheet.append(["PLACA", "MARCA", "MODELO"])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    headers = {
        "Content-Disposition": "attachment; filename=plantilla_vehiculos.xlsx"
    }
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )


# ============================================
# ACTUALIZAR VEHÍCULO
# ============================================

@router.put("/{vehiculo_id}", response_model=VehiculoResponse)
def actualizar_vehiculo(
    vehiculo_id: int,
    vehiculo_update: VehiculoUpdate,
    current_user: Usuario = Depends(require_permission("editar_vehiculo")),
    db: Session = Depends(get_db)
):
    """
    Actualiza un vehículo.
    
    PUT /vehiculos/1
    Body:
    {
        "estado": "inactivo"
    }
    """
    
    vehiculo = db.query(Vehiculo).filter(Vehiculo.id == vehiculo_id).first()

    if not vehiculo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Vehículo no encontrado"
        )

    # Multi-tenancy: verificar pertenencia a empresa excepto super_admin
    if current_user.rol and current_user.rol.nombre != "super_admin":
        if vehiculo.empresa_id != current_user.empresa_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="No tienes permiso para acceder a este recurso"
            )

    # Actualizar solo los campos enviados
    for campo, valor in vehiculo_update.model_dump(exclude_unset=True).items():
        setattr(vehiculo, campo, valor)

    db.add(vehiculo)
    db.commit()
    db.refresh(vehiculo)

    return vehiculo


# ============================================
# ELIMINAR VEHÍCULO
# ============================================

@router.delete("/{vehiculo_id}", status_code=status.HTTP_204_NO_CONTENT)
def eliminar_vehiculo(
    vehiculo_id: int,
    current_user: Usuario = Depends(require_permission("eliminar_vehiculo")),
    db: Session = Depends(get_db)
):
    """
    Marca un vehículo como INACTIVO (soft delete).
    
    DELETE /vehiculos/1
    
    No elimina los datos, los marca como inactivos. Preserva la auditoría
    y permite recuperación si es necesario.
    """
    
    vehiculo = db.query(Vehiculo).filter(Vehiculo.id == vehiculo_id).first()

    if not vehiculo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Vehículo no encontrado"
        )

    # Multi-tenancy: verificar pertenencia a empresa excepto super_admin
    if current_user.rol and current_user.rol.nombre != "super_admin":
        if vehiculo.empresa_id != current_user.empresa_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="No tienes permiso para acceder a este recurso"
            )

    # Soft Delete: cambiar estado a inactivo
    vehiculo.estado = EstadoGeneral.inactivo
    db.add(vehiculo)
    db.commit()

    return None
